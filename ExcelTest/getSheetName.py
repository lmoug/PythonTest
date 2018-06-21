#  -*- coding:gbk -*-  
  
import openpyxl  
  

wb1 = openpyxl.load_workbook('E:\工作-项目文档\博智项目\上线准备\原始文档\众信泰富电商一期项目_数据_2B中台数据收集模板_V1.0_180111.xlsx')  
wb2=openpyxl.load_workbook(r'E:\工作-项目文档\博智项目\上线准备\原始文档\众信泰富电商一期项目_数据_2C中台数据收集模板_V1.0_180111.xlsx')
wb3=openpyxl.load_workbook(r'E:\工作-项目文档\博智项目\上线准备\原始文档\众信泰富电商一期项目_数据_WMS(2B)数据收集模板_V1.0_180111.xlsx')
wb4= openpyxl.load_workbook(r'E:\工作-项目文档\博智项目\上线准备\原始文档\众信泰富电商一期项目_数据_WMS(2C)数据收集模板_V1.0_180111.xlsx')
  
#获取workbook中所有的表格  
print ("//////////////////////////////////2B中台数据收集模板")
sheets = wb1.get_sheet_names()  
  
for i in sheets:
    print i
    
print ("//////////////////////////////////2C中台数据收集模板")
sheets = wb2.get_sheet_names()  
  
for i in sheets:
    print i
    
print ("//////////////////////////////////WMS(2B)数据收集模板")
sheets = wb3.get_sheet_names()  
  
for i in sheets:
    print i 
    
print ("//////////////////////////////////WMS(2C)数据收集模板")
sheets = wb4.get_sheet_names()  
  
for i in sheets:
    print i    
#循环遍历所有sheet  
#for i in range(len(sheets)):  
#    sheet= wb.get_sheet_by_name(sheets[i])  
      
#   print('\n\n第'+str(i+1)+'个sheet: ' + sheet.title+'->>>')  
  
#    for r in range(1,sheet.max_row+1):  
#        if r == 1:  
#            print('\n'+''.join([str(sheet.cell(row=r,column=c).value).ljust(17) for c in range(1,sheet.max_column+1)] ))  
#        else:  
#            print(''.join([str(sheet.cell(row=r,column=c).value).ljust(20) for c in range(1,sheet.max_column+1)] ))  